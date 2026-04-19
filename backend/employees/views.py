from rest_framework import viewsets, status, generics, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
from django.db.models import Q, Count
from django_filters.rest_framework import DjangoFilterBackend
from datetime import datetime, date
import openpyxl
import csv
from io import BytesIO
from django.http import HttpResponse
from collections import defaultdict

from .models import User, Employee, Team, VPIndia, ReportingManager
from .serializers import (
    UserSerializer, EmployeeSerializer, TeamSerializer,
    VPIndiaSerializer, ReportingManagerSerializer,
    DashboardStatsSerializer, ExitTrendSerializer
)
from .permissions import IsAdminUser

@api_view(['POST'])
@permission_classes([AllowAny])
def signup(request):
    serializer = UserSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        refresh = RefreshToken.for_user(user)
        return Response({
            'user': UserSerializer(user).data,
            'refresh': str(refresh),
            'access': str(refresh.access_token),
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    username = request.data.get('username')
    password = request.data.get('password')
    
    user = authenticate(username=username, password=password)
    if user:
        refresh = RefreshToken.for_user(user)
        return Response({
            'user': UserSerializer(user).data,
            'refresh': str(refresh),
            'access': str(refresh.access_token),
        })
    return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

@api_view(['POST'])
@permission_classes([AllowAny])
def password_reset_request(request):
    email = request.data.get('email')
    try:
        user = User.objects.get(email=email)
        return Response({'message': 'Password reset link sent to your email'})
    except User.DoesNotExist:
        return Response({'error': 'User with this email does not exist'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
@permission_classes([AllowAny])
def password_reset_confirm(request):
    email = request.data.get('email')
    new_password = request.data.get('new_password')
    
    try:
        user = User.objects.get(email=email)
        user.set_password(new_password)
        user.save()
        return Response({'message': 'Password reset successful'})
    except User.DoesNotExist:
        return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    active_employees = Employee.objects.filter(status='Active')
    
    stats = {
        'full_time_count': active_employees.filter(type='Full Time').count(),
        'intern_count': active_employees.filter(type='Intern').count(),
        'contract_count': active_employees.filter(type='Contract').count(),
        'total_count': active_employees.count(),
    }
    
    serializer = DashboardStatsSerializer(stats)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exit_trends(request):
    year = request.query_params.get('year', date.today().year)
    
    exited_employees = Employee.objects.filter(
        status='Exited',
        exit_date__year=year
    )
    
    trends = defaultdict(lambda: {'voluntary': 0, 'termination': 0})
    
    for emp in exited_employees:
        if emp.exit_quarter:
            quarter = emp.exit_quarter
            if emp.exit_type == 'Voluntary':
                trends[quarter]['voluntary'] += 1
            elif emp.exit_type == 'Termination':
                trends[quarter]['termination'] += 1
    
    result = []
    for q in range(1, 5):
        quarter_key = f"{year} Q{q}"
        voluntary = trends[quarter_key]['voluntary']
        termination = trends[quarter_key]['termination']
        result.append({
            'quarter': quarter_key,
            'voluntary_exits': voluntary,
            'terminations': termination,
            'total_exits': voluntary + termination,
            'difference': voluntary - termination
        })
    
    serializer = ExitTrendSerializer(result, many=True)
    return Response(serializer.data)

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['type', 'status', 'team']
    search_fields = ['employee_id', 'name', 'email', 'title', 'reporting_to', 'vp_india']
    ordering_fields = ['date_of_joining', 'name', 'employee_id']
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        return [IsAuthenticated()]
    
    @action(detail=False, methods=['post'], permission_classes=[IsAdminUser])
    def import_employees(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                rows = list(sheet.iter_rows(values_only=True))
                headers = rows[0]
                data_rows = rows[1:]
            elif file.name.endswith('.csv'):
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.reader(decoded_file)
                headers = next(reader)
                data_rows = list(reader)
            else:
                return Response({'error': 'Invalid file format'}, status=status.HTTP_400_BAD_REQUEST)
            
            created_count = 0
            errors = []
            
            for idx, row in enumerate(data_rows, start=2):
                try:
                    team_name = row[headers.index('team')]
                    team, _ = Team.objects.get_or_create(
                        team_name=team_name,
                        defaults={
                            'us_team_head': row[headers.index('us_team_head')] if 'us_team_head' in headers else '',
                            'india_head': row[headers.index('india_head')] if 'india_head' in headers else ''
                        }
                    )
                    
                    employee_data = {
                        'employee_id': row[headers.index('employee_id')],
                        'title': row[headers.index('title')],
                        'name': row[headers.index('name')],
                        'email': row[headers.index('email')],
                        'date_of_joining': datetime.strptime(row[headers.index('date_of_joining')], '%Y-%m-%d').date(),
                        'type': row[headers.index('type')],
                        'status': row[headers.index('status')],
                        'reporting_to': row[headers.index('reporting_to')],
                        'team': team,
                        'vp_india': row[headers.index('vp_india')],
                        'experience_prior_adf': float(row[headers.index('experience_prior_adf')]),
                    }
                    
                    if 'exit_date' in headers and row[headers.index('exit_date')]:
                        employee_data['exit_date'] = datetime.strptime(row[headers.index('exit_date')], '%Y-%m-%d').date()
                    if 'exit_type' in headers and row[headers.index('exit_type')]:
                        employee_data['exit_type'] = row[headers.index('exit_type')]
                    
                    Employee.objects.create(**employee_data)
                    created_count += 1
                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")
            
            return Response({
                'message': f'Successfully imported {created_count} employees',
                'errors': errors
            })
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'], url_path='advanced-search')
    def advanced_search(self, request):
        queryset = Employee.objects.all()
        
        # Text-based searches - exact matches only
        employee_id = request.query_params.get('employee_id', '')
        name = request.query_params.get('name', '')
        reporting_to = request.query_params.get('reporting_to', '')
        vp_india = request.query_params.get('vp_india', '')
        
        if employee_id:
            queryset = queryset.filter(employee_id__iexact=employee_id)
        if name:
            queryset = queryset.filter(name__iexact=name)
        if reporting_to:
            queryset = queryset.filter(reporting_to__iexact=reporting_to)
        if vp_india:
            queryset = queryset.filter(vp_india__iexact=vp_india)
        
        # Single DOJ search
        doj = request.query_params.get('doj')
        
        if doj:
            try:
                doj_date = datetime.strptime(doj, '%Y-%m-%d').date()
                queryset = queryset.filter(date_of_joining=doj_date)
            except ValueError:
                pass
        
        # Tenure search - exact match
        tenure = request.query_params.get('tenure')
        if tenure:
            try:
                # Normalize the tenure input to match the format stored in tenure_at_adf property
                tenure_normalized = tenure.strip().lower()
                
                # Parse tenure input to normalize format (e.g., "2y 3m" -> "2y 3m", "2y" -> "2y", "6m" -> "6m")
                tenure_parts = tenure_normalized.split()
                years = 0
                months = 0
                
                for part in tenure_parts:
                    if part.endswith('y'):
                        years = int(part[:-1])
                    elif part.endswith('m'):
                        months = int(part[:-1])
                
                # Create the exact format that tenure_at_adf property returns
                if years > 0 and months > 0:
                    exact_tenure = f"{years}y {months}m"
                elif years > 0:
                    exact_tenure = f"{years}y"
                elif months > 0:
                    exact_tenure = f"{months}m"
                else:
                    # Invalid format, skip filtering
                    pass
                
                # Filter employees by exact tenure match using Python evaluation
                # This ensures exact matching with the tenure_at_adf property
                employee_ids = []
                for emp in queryset:
                    if emp.tenure_at_adf == exact_tenure:
                        employee_ids.append(emp.id)
                
                queryset = queryset.filter(id__in=employee_ids)
                
            except (ValueError, AttributeError):
                pass
        
        # Team filtering
        team_id = request.query_params.get('team')
        if team_id:
            queryset = queryset.filter(team_id=team_id)
        
        # Additional filters
        type_filter = request.query_params.get('type')
        status_filter = request.query_params.get('status')
        
        if type_filter:
            queryset = queryset.filter(type=type_filter)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Ordering
        ordering = request.query_params.get('ordering', '-date_of_joining')
        queryset = queryset.order_by(ordering)
        
        # Pagination
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        
        from django.core.paginator import Paginator
        paginator = Paginator(queryset, page_size)
        page_obj = paginator.get_page(page)
        
        serializer = self.get_serializer(page_obj.object_list, many=True)
        
        return Response({
            'results': serializer.data,
            'count': paginator.count,
            'next': page_obj.has_next() and page_obj.next_page_number() or None,
            'previous': page_obj.has_previous() and page_obj.previous_page_number() or None,
            'current_page': page,
            'total_pages': paginator.num_pages,
        })
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        format_type = request.query_params.get('format', 'csv')
        queryset = self.filter_queryset(self.get_queryset())
        
        if format_type == 'xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Employees"
            
            headers = ['Emp ID', 'Title', 'Name', 'Email', 'DOJ', 
                      'Tenure at ADF', 'Experience Prior', 'Type', 'Status', 
                      'Reporting To', 'Team', 'VP India']
            ws.append(headers)
            
            for emp in queryset:
                ws.append([
                    emp.employee_id, emp.title, emp.name, emp.email,
                    emp.date_of_joining.strftime('%Y-%m-%d'), emp.tenure_at_adf,
                    float(emp.experience_prior_adf), emp.type, emp.status,
                    emp.reporting_to, emp.team.team_name if emp.team else '',
                    emp.vp_india
                ])
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=employees_{datetime.now().strftime("%Y%m%d")}.xlsx'
            return response
        else:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename=employees_{datetime.now().strftime("%Y%m%d")}.csv'
            
            writer = csv.writer(response)
            writer.writerow(['Emp ID', 'Title', 'Name', 'Email', 'DOJ', 
                           'Tenure at ADF', 'Experience Prior', 'Type', 'Status', 
                           'Reporting To', 'Team', 'VP India'])
            
            for emp in queryset:
                writer.writerow([
                    emp.employee_id, emp.title, emp.name, emp.email,
                    emp.date_of_joining.strftime('%Y-%m-%d'), emp.tenure_at_adf,
                    float(emp.experience_prior_adf), emp.type, emp.status,
                    emp.reporting_to, emp.team.team_name if emp.team else '',
                    emp.vp_india
                ])
            
            return response

class TeamViewSet(viewsets.ModelViewSet):
    queryset = Team.objects.all()
    serializer_class = TeamSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['team_name', 'us_team_head', 'india_head']
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        return [IsAuthenticated()]
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        queryset = self.get_queryset()
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=teams_{datetime.now().strftime("%Y%m%d")}.csv'
        
        writer = csv.writer(response)
        writer.writerow(['US Team Head', 'India Head', 'Team Name', 'Employee Count'])
        
        for team in queryset:
            writer.writerow([
                team.us_team_head, team.india_head, team.team_name,
                team.employees.filter(status='Active').count()
            ])
        
        return response

class VPIndiaViewSet(viewsets.ModelViewSet):
    queryset = VPIndia.objects.all()
    serializer_class = VPIndiaSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        return [IsAuthenticated()]

class ReportingManagerViewSet(viewsets.ModelViewSet):
    queryset = ReportingManager.objects.all()
    serializer_class = ReportingManagerSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        return [IsAuthenticated()]
